import openpyxl
def crearexcel():
    wb = openpyxl.Workbook()
    ws = wb.active
    header = ["Nombre", "Tipo 1", "Tipo 2", "Altura", "Peso", None, "Promedio de peso", None, "Promedio de altura", None, "Moda de tipos"]
    ws.append(header)
    ws.title = "Consulta"
    wb.save("pokemon.xlsx")

def agregarvalores(valores):
    wb = openpyxl.load_workbook("pokemon.xlsx")
    ws = wb.active
    ws.append(valores)
    wb.save("pokemon.xlsx")

def agregarcalculos(columna, valor):
    wb = openpyxl.load_workbook("pokemon.xlsx")
    ws = wb.active
    column_index = columna
    max_row = ws.max_row
    valoragregado = valor
    ws.cell(row=max_row + 1, column=column_index).value = valoragregado
    wb.save("pokemon.xlsx")